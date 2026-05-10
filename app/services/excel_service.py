import re
import uuid
from pathlib import Path
from typing import Any

from fastapi import HTTPException, UploadFile
from openpyxl import Workbook, load_workbook
from openpyxl.utils import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from app.core.config import settings

_SAFE_NAME_RE = re.compile(r"[^a-zA-Z0-9._-]+")


def sanitize_filename(filename: str) -> str:
    filename = Path(filename).name.strip() or "workbook.xlsx"
    filename = _SAFE_NAME_RE.sub("_", filename)
    if not filename.lower().endswith(".xlsx"):
        filename += ".xlsx"
    return filename


def new_workbook_id() -> str:
    return uuid.uuid4().hex


def workbook_path(workbook_id: str) -> Path:
    if not re.fullmatch(r"[a-f0-9]{32}", workbook_id):
        raise HTTPException(status_code=400, detail="Invalid workbook_id")
    path = settings.storage_dir / f"{workbook_id}.xlsx"
    return path


def require_workbook_path(workbook_id: str) -> Path:
    path = workbook_path(workbook_id)
    if not path.exists():
        raise HTTPException(status_code=404, detail="Workbook not found")
    return path


def metadata_path(workbook_id: str) -> Path:
    return settings.storage_dir / f"{workbook_id}.name"


def save_filename(workbook_id: str, filename: str) -> None:
    metadata_path(workbook_id).write_text(sanitize_filename(filename), encoding="utf-8")


def read_filename(workbook_id: str) -> str:
    meta = metadata_path(workbook_id)
    if meta.exists():
        return meta.read_text(encoding="utf-8").strip() or f"{workbook_id}.xlsx"
    return f"{workbook_id}.xlsx"


def get_sheet_or_404(wb, sheet_name: str) -> Worksheet:
    if sheet_name not in wb.sheetnames:
        raise HTTPException(status_code=404, detail=f"Sheet '{sheet_name}' not found")
    return wb[sheet_name]


def create_workbook(filename: str, sheet_name: str, headers: list[str] | None, rows: list[list[Any]]) -> tuple[str, str]:
    workbook_id = new_workbook_id()
    clean_name = sanitize_filename(filename)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    if headers:
        ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(workbook_path(workbook_id))
    save_filename(workbook_id, clean_name)
    return workbook_id, clean_name


async def upload_workbook(file: UploadFile) -> tuple[str, str]:
    clean_name = sanitize_filename(file.filename or "uploaded.xlsx")
    if not clean_name.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")

    raw = await file.read()
    max_bytes = settings.max_upload_mb * 1024 * 1024
    if len(raw) > max_bytes:
        raise HTTPException(status_code=413, detail=f"File too large. Max {settings.max_upload_mb} MB")

    workbook_id = new_workbook_id()
    path = workbook_path(workbook_id)
    path.write_bytes(raw)

    try:
        wb = load_workbook(path)
        wb.close()
    except Exception as exc:
        path.unlink(missing_ok=True)
        raise HTTPException(status_code=400, detail=f"Invalid .xlsx file: {exc}") from exc

    save_filename(workbook_id, clean_name)
    return workbook_id, clean_name


def list_workbooks() -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    for path in sorted(settings.storage_dir.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True):
        workbook_id = path.stem
        stat = path.stat()
        items.append({
            "workbook_id": workbook_id,
            "filename": read_filename(workbook_id),
            "size_bytes": stat.st_size,
            "modified_at": stat.st_mtime,
        })
    return items


def delete_workbook(workbook_id: str) -> None:
    path = require_workbook_path(workbook_id)
    path.unlink(missing_ok=True)
    metadata_path(workbook_id).unlink(missing_ok=True)


def list_sheets(workbook_id: str) -> list[str]:
    path = require_workbook_path(workbook_id)
    wb = load_workbook(path)
    names = wb.sheetnames
    wb.close()
    return names


def create_sheet(workbook_id: str, sheet_name: str) -> list[str]:
    path = require_workbook_path(workbook_id)
    wb = load_workbook(path)
    if sheet_name in wb.sheetnames:
        wb.close()
        raise HTTPException(status_code=409, detail="Sheet already exists")
    wb.create_sheet(sheet_name)
    wb.save(path)
    names = wb.sheetnames
    wb.close()
    return names


def rename_sheet(workbook_id: str, sheet_name: str, new_sheet_name: str) -> list[str]:
    path = require_workbook_path(workbook_id)
    wb = load_workbook(path)
    ws = get_sheet_or_404(wb, sheet_name)
    if new_sheet_name in wb.sheetnames:
        wb.close()
        raise HTTPException(status_code=409, detail="New sheet name already exists")
    ws.title = new_sheet_name
    wb.save(path)
    names = wb.sheetnames
    wb.close()
    return names


def delete_sheet(workbook_id: str, sheet_name: str) -> list[str]:
    path = require_workbook_path(workbook_id)
    wb = load_workbook(path)
    if len(wb.sheetnames) <= 1:
        wb.close()
        raise HTTPException(status_code=400, detail="Cannot delete the last sheet")
    ws = get_sheet_or_404(wb, sheet_name)
    wb.remove(ws)
    wb.save(path)
    names = wb.sheetnames
    wb.close()
    return names


def read_rows(workbook_id: str, sheet_name: str, start_row: int, limit: int) -> list[list[Any]]:
    if start_row < 1 or limit < 1 or limit > 1000:
        raise HTTPException(status_code=400, detail="start_row must be >= 1 and limit must be 1..1000")
    path = require_workbook_path(workbook_id)
    wb = load_workbook(path, data_only=False)
    ws = get_sheet_or_404(wb, sheet_name)
    end_row = min(ws.max_row, start_row + limit - 1)
    rows = [[cell.value for cell in row] for row in ws.iter_rows(min_row=start_row, max_row=end_row)]
    wb.close()
    return rows


def append_row(workbook_id: str, sheet_name: str, values: list[Any]) -> int:
    path = require_workbook_path(workbook_id)
    wb = load_workbook(path)
    ws = get_sheet_or_404(wb, sheet_name)
    ws.append(values)
    row_index = ws.max_row
    wb.save(path)
    wb.close()
    return row_index


def update_row(workbook_id: str, sheet_name: str, row_index: int, values: list[Any]) -> None:
    if row_index < 1:
        raise HTTPException(status_code=400, detail="row_index must be >= 1")
    path = require_workbook_path(workbook_id)
    wb = load_workbook(path)
    ws = get_sheet_or_404(wb, sheet_name)
    for col_idx, value in enumerate(values, start=1):
        ws.cell(row=row_index, column=col_idx, value=value)
    wb.save(path)
    wb.close()


def delete_row(workbook_id: str, sheet_name: str, row_index: int) -> None:
    if row_index < 1:
        raise HTTPException(status_code=400, detail="row_index must be >= 1")
    path = require_workbook_path(workbook_id)
    wb = load_workbook(path)
    ws = get_sheet_or_404(wb, sheet_name)
    if row_index > ws.max_row:
        wb.close()
        raise HTTPException(status_code=404, detail="Row not found")
    ws.delete_rows(row_index, 1)
    wb.save(path)
    wb.close()


def read_cell(workbook_id: str, sheet_name: str, cell: str) -> Any:
    path = require_workbook_path(workbook_id)
    wb = load_workbook(path, data_only=False)
    ws = get_sheet_or_404(wb, sheet_name)
    value = ws[cell].value
    wb.close()
    return value


def update_cell(workbook_id: str, sheet_name: str, cell: str, value: Any) -> None:
    path = require_workbook_path(workbook_id)
    wb = load_workbook(path)
    ws = get_sheet_or_404(wb, sheet_name)
    ws[cell] = value
    wb.save(path)
    wb.close()


def read_range(workbook_id: str, sheet_name: str, range_ref: str) -> list[list[Any]]:
    path = require_workbook_path(workbook_id)
    wb = load_workbook(path, data_only=False)
    ws = get_sheet_or_404(wb, sheet_name)
    try:
        min_col, min_row, max_col, max_row = range_boundaries(range_ref)
    except ValueError as exc:
        wb.close()
        raise HTTPException(status_code=400, detail="Invalid range. Example: A1:D10") from exc
    values = []
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        values.append([cell.value for cell in row])
    wb.close()
    return values


def write_range(workbook_id: str, sheet_name: str, start_cell: str, values: list[list[Any]]) -> None:
    path = require_workbook_path(workbook_id)
    wb = load_workbook(path)
    ws = get_sheet_or_404(wb, sheet_name)
    try:
        min_col, min_row, _, _ = range_boundaries(f"{start_cell}:{start_cell}")
    except ValueError as exc:
        wb.close()
        raise HTTPException(status_code=400, detail="Invalid start_cell. Example: A1") from exc
    for r_offset, row in enumerate(values):
        for c_offset, value in enumerate(row):
            ws.cell(row=min_row + r_offset, column=min_col + c_offset, value=value)
    wb.save(path)
    wb.close()
